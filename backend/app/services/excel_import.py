from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import InventoryLabel
from app.utils.excel_parser import (
    find_column_mapping,
    parse_label,
    parse_quantity,
    parse_reference,
    parse_fifo_datetime,
    format_fifo_datetime,
    REQUIRED_COLUMNS,
)


@dataclass
class RowError:
    row: int
    reason: str
    error_type: str = "general"


@dataclass
class ImportPreviewRow:
    label: str
    reference: str
    quantity: float
    fifo_date: str


@dataclass
class ImportResult:
    total_rows: int = 0
    successful: int = 0
    errors: list[RowError] = field(default_factory=list)
    duplicate_labels: list[str] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)
    invalid_label_count: int = 0
    invalid_reference_count: int = 0
    invalid_quantity_count: int = 0
    invalid_fifo_date_count: int = 0
    preview_rows: list[ImportPreviewRow] = field(default_factory=list)


def import_excel(db: Session, file_content: bytes, replace_existing: bool = True) -> ImportResult:
    result = ImportResult()

    try:
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception:
        result.errors.append(RowError(row=0, reason="Excel dosyası açılamadı"))
        return result

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result.errors.append(RowError(row=0, reason="Excel dosyası boş"))
        return result

    headers = rows[0]
    mapping = find_column_mapping(list(headers))

    missing = []
    for req_key, field_name in REQUIRED_COLUMNS.items():
        if field_name not in mapping:
            missing.append(req_key)
    if missing:
        result.missing_columns = missing
        return result

    if replace_existing:
        db.query(InventoryLabel).delete()
        db.commit()

    seen_labels: set[str] = set()
    batch: list[InventoryLabel] = []
    BATCH_SIZE = 1000
    PREVIEW_LIMIT = 5

    for row_idx, row in enumerate(rows[1:], start=2):
        result.total_rows += 1

        label_val = row[mapping["label"]] if mapping["label"] < len(row) else None
        ref_val = row[mapping["reference"]] if mapping["reference"] < len(row) else None
        qty_val = row[mapping["quantity"]] if mapping["quantity"] < len(row) else None
        date_val = row[mapping["fifo_date"]] if mapping["fifo_date"] < len(row) else None

        label = parse_label(label_val)
        if not label:
            result.invalid_label_count += 1
            result.errors.append(RowError(row=row_idx, reason="ETİKET boş veya geçersiz", error_type="label"))
            continue

        reference = parse_reference(ref_val)
        if not reference:
            result.invalid_reference_count += 1
            result.errors.append(RowError(row=row_idx, reason="REFERANS boş veya geçersiz", error_type="reference"))
            continue

        quantity = parse_quantity(qty_val)
        if quantity is None:
            result.invalid_quantity_count += 1
            result.errors.append(RowError(row=row_idx, reason="MİKTAR geçersiz", error_type="quantity"))
            continue

        fifo_date = parse_fifo_datetime(date_val)
        if fifo_date is None:
            result.invalid_fifo_date_count += 1
            result.errors.append(RowError(row=row_idx, reason="X98FIFO_TARIH geçersiz", error_type="fifo_date"))
            continue

        if label in seen_labels:
            result.duplicate_labels.append(label)
            result.errors.append(RowError(row=row_idx, reason=f"Duplicate etiket: {label}", error_type="duplicate"))
            continue

        seen_labels.add(label)
        batch.append(InventoryLabel(
            label=label,
            reference=reference,
            quantity=quantity,
            fifo_date=fifo_date,
        ))

        if len(result.preview_rows) < PREVIEW_LIMIT:
            result.preview_rows.append(ImportPreviewRow(
                label=label,
                reference=reference,
                quantity=float(quantity),
                fifo_date=format_fifo_datetime(fifo_date),
            ))

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.commit()
            result.successful += len(batch)
            batch = []

    if batch:
        db.bulk_save_objects(batch)
        db.commit()
        result.successful += len(batch)

    wb.close()
    return result
