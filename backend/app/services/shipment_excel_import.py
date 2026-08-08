from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.target_service import add_target
from app.utils.excel_parser import (
    find_shipment_column_mapping,
    parse_reference,
    parse_quantity,
)


@dataclass
class ShipmentRowError:
    row: int
    reason: str


@dataclass
class ShipmentTargetImportResult:
    total_rows: int = 0
    successful: int = 0
    errors: list[ShipmentRowError] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)
    targets: list[dict] = field(default_factory=list)


def import_shipment_targets_excel(
    db: Session,
    file_content: bytes,
    replace: bool = True,
) -> ShipmentTargetImportResult:
    """Sevkiyat Excelini hedef listesine aktarır (FIFO henüz hesaplanmaz)."""
    from app.models import ShipmentTarget

    result = ShipmentTargetImportResult()

    try:
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception:
        result.errors.append(ShipmentRowError(row=0, reason="Excel dosyası açılamadı"))
        return result

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result.errors.append(ShipmentRowError(row=0, reason="Excel dosyası boş"))
        return result

    mapping = find_shipment_column_mapping(list(rows[0]))
    if "reference" not in mapping:
        result.missing_columns.append("REFERANS")
    if "quantity" not in mapping:
        result.missing_columns.append("MİKTAR")
    if result.missing_columns:
        wb.close()
        return result

    if replace:
        db.query(ShipmentTarget).delete()
        db.commit()

    for row_idx, row in enumerate(rows[1:], start=2):
        result.total_rows += 1
        ref_val = row[mapping["reference"]] if mapping["reference"] < len(row) else None
        qty_val = row[mapping["quantity"]] if mapping["quantity"] < len(row) else None

        reference = parse_reference(ref_val)
        if not reference:
            result.errors.append(ShipmentRowError(row=row_idx, reason="REFERANS boş"))
            continue

        quantity = parse_quantity(qty_val)
        if quantity is None:
            result.errors.append(ShipmentRowError(row=row_idx, reason="MİKTAR geçersiz"))
            continue

        try:
            t = add_target(db, reference, quantity)
            result.successful += 1
            result.targets.append({
                "id": t.id,
                "reference": t.reference,
                "target_quantity": t.target_quantity,
            })
        except ValueError as e:
            result.errors.append(ShipmentRowError(row=row_idx, reason=str(e)))

    wb.close()
    return result
